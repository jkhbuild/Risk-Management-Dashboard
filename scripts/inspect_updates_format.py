"""Read-only inspection of Risk_Updates conventions to inform regeneration.
Prints: distinct authors and counts; update_id format and range; how year
disambiguation was applied (year sequence per risk vs date); examples of
mitigation_log raw text to spot edge cases in the splitter regex.

Run: python scripts/inspect_updates_format.py
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
REGISTER_PATH = ROOT / "source_data" / "Tonnelle_Risk_Register_260519.xlsx"
UPDATES_PATH = ROOT / "source_data" / "Tonnelle_Risk_Updates_MASTER.xlsx"


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def main():
    print("=" * 72)
    print("UPDATES: author / id / year conventions")
    print("=" * 72)
    wb = load_workbook(UPDATES_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    print(f"row count: {len(rows)}")
    print(f"headers: {headers}")

    print("\n-- distinct authors")
    auth = Counter(r.get("author") for r in rows)
    for a, n in sorted(auth.items(), key=lambda kv: (-kv[1], str(kv[0]))):
        print(f"   {a!r}: {n}")

    print("\n-- update_id format")
    ids = [r.get("update_id") for r in rows]
    types = Counter(type(i).__name__ for i in ids)
    print(f"   types: {dict(types)}")
    print(f"   min={min(ids)}  max={max(ids)}  span={max(ids)-min(ids)+1}")
    print(f"   strictly sequential? {ids == list(range(min(ids), max(ids)+1))}")

    print("\n-- update_year column vs update_date.year")
    mismatches = []
    for r in rows:
        stored = r.get("update_year")
        d = to_date(r.get("update_date"))
        if d and stored != d.year:
            mismatches.append((r.get("update_id"), stored, d))
    print(f"   mismatches: {len(mismatches)}")
    for m in mismatches[:5]:
        print(f"     {m}")

    print("\n-- year sequence per risk (does year flip mid-risk? identifies"
          " how the splitter ordered M/D within a year)")
    by_risk = defaultdict(list)
    for r in rows:
        d = to_date(r.get("update_date"))
        if d:
            by_risk[r.get("risk_id")].append(d)
    for rid in sorted(by_risk.keys())[:6]:
        dts = sorted(by_risk[rid])
        years = [d.year for d in dts]
        print(f"   {rid}: {dts}")
        flips = sum(1 for i in range(1, len(years)) if years[i] != years[i-1])
        if flips:
            print(f"      year flips: {flips}")

    print("\n" + "=" * 72)
    print("REGISTER: mitigation_log raw samples (first 3 risks)")
    print("=" * 72)
    wb2 = load_workbook(REGISTER_PATH, data_only=True)
    ws2 = wb2["Risk_Register"]
    hdrs2 = [c.value for c in ws2[1]]
    reg_rows = []
    for r in ws2.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        reg_rows.append(dict(zip(hdrs2, r)))
    for r in reg_rows[:3]:
        print(f"\n-- {r.get('risk_id')}")
        log = r.get("mitigation_log") or ""
        print(f"raw length: {len(log)}")
        print(f"raw text: {log!r}")

    print("\n-- mitigation_log dated-entry counts per risk")
    DATE_LEADER = re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})\s*-\s*", re.MULTILINE)
    total = 0
    for r in reg_rows:
        log = r.get("mitigation_log") or ""
        n = len(DATE_LEADER.findall(log))
        total += n
        print(f"   {r.get('risk_id')}: {n} dated entries")
    print(f"\n   TOTAL dated entries across all 37 risks: {total}")


if __name__ == "__main__":
    main()
