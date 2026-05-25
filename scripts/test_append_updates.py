"""Unit tests for scripts/append_updates.py.

Run from project root:
  PYTHONUTF8=1 PYTHONIOENCODING=utf-8 python -m pytest scripts/test_append_updates.py -v
"""
from __future__ import annotations

import logging
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import append_updates as au  # noqa: E402


# ---------------------------------------------------------------- fixtures


def make_register_xlsx(path: Path, rows: list[dict]) -> None:
    """Write a minimal Risk_Register .xlsx with only the columns the script reads."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk_Register"
    cols = ["risk_id", "risk_coordinator", "mitigation_log"]
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    wb.save(path)


def make_updates_xlsx(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk_Updates"
    ws.append(au.UPDATE_COLS)
    for r in rows:
        ws.append([r.get(c) for c in au.UPDATE_COLS])
    wb.save(path)


@pytest.fixture
def empty_updates(tmp_path: Path) -> Path:
    p = tmp_path / "updates.xlsx"
    make_updates_xlsx(p, [])
    return p


# ---------------------------------------------------------------- parse_log


def test_parse_log_two_entries():
    log = "5/10 - first event. 5/22 - second event with detail."
    out = au.parse_log(log)
    assert out == [
        (5, 10, "first event."),
        (5, 22, "second event with detail."),
    ]


def test_parse_log_empty_or_none():
    assert au.parse_log("") == []
    assert au.parse_log(None) == []
    assert au.parse_log(123) == []  # non-string


def test_parse_log_no_dated_markers():
    assert au.parse_log("just some prose with no date leaders") == []


def test_parse_log_malformed_does_not_crash(caplog):
    """Out-of-range M/D logs a warning and is skipped; valid entries still parsed."""
    log = "13/45 - invalid date entry. 5/22 - valid entry."
    with caplog.at_level(logging.WARNING):
        out = au.parse_log(log, source_rid="TONN-CON.TEST")
    assert out == [(5, 22, "valid entry.")]
    assert any("out-of-range M/D" in r.message or "13/45" in r.message
               for r in caplog.records)


def test_parse_log_empty_body_logged_and_skipped(caplog):
    log = "5/22 - 5/23 - real entry."
    with caplog.at_level(logging.WARNING):
        out = au.parse_log(log)
    assert out == [(5, 23, "real entry.")]
    assert any("empty body" in r.message for r in caplog.records)


# ---------------------------------------------------------------- year inference


def test_infer_year_default_current_year():
    today = date(2026, 5, 23)
    assert au.infer_year(5, 22, today, None) == 2026


def test_infer_year_more_than_six_months_future_rolls_back():
    """5/23/2026 + a 12/1 entry: 2026-12-01 is ~6.3 months out, roll back to 2025."""
    today = date(2026, 5, 23)
    assert au.infer_year(12, 1, today, None) == 2025


def test_infer_year_just_under_six_months_stays_current():
    """5/23/2026 + 11/15 = ~176d out, < 183, stays 2026."""
    today = date(2026, 5, 23)
    assert au.infer_year(11, 15, today, None) == 2026


def test_infer_year_past_date_stays_current():
    today = date(2026, 5, 23)
    assert au.infer_year(1, 5, today, None) == 2026


def test_infer_year_override_wins():
    today = date(2026, 5, 23)
    assert au.infer_year(5, 22, today, override=2024) == 2024


def test_infer_year_override_invalid_returns_none(caplog):
    today = date(2026, 5, 23)
    with caplog.at_level(logging.WARNING):
        assert au.infer_year(2, 30, today, override=2025) is None


def test_infer_year_feb29_non_leap_returns_none(caplog):
    today = date(2025, 5, 23)  # 2025 not leap
    with caplog.at_level(logging.WARNING):
        assert au.infer_year(2, 29, today, None) is None


# ---------------------------------------------------------------- compute_appends


def test_compute_appends_detects_one_new(tmp_path: Path):
    """One Register risk with two log entries; one already in Updates -> one append."""
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/10 - first event. 5/22 - new event.",
    }]
    existing = [{
        "update_id": 1,
        "risk_id": "TONN-CON.01",
        "update_date": date(2026, 5, 10),
        "update_year": 2026,
        "author": "Anton Benedict",
        "note": "first event.",
    }]
    new_rows = au.compute_appends(register, existing, today, None, None)
    assert len(new_rows) == 1
    row = new_rows[0]
    assert row["risk_id"] == "TONN-CON.01"
    assert row["update_date"] == date(2026, 5, 22)
    assert row["update_year"] == 2026
    assert row["author"] == "Anton Benedict"
    assert row["note"] == "new event."
    assert row["update_id"] == 2  # max_id + 1


def test_compute_appends_idempotent(tmp_path: Path):
    """Running twice against the same state yields the same result both times."""
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/10 - event A. 5/22 - event B.",
    }]
    existing = [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2026, 5, 10), "update_year": 2026,
        "author": "Anton Benedict", "note": "event A.",
    }]
    pass1 = au.compute_appends(register, existing, today, None, None)
    assert len(pass1) == 1
    # Simulate: pass1's rows are now in existing
    combined = existing + [
        {**r, "update_date": r["update_date"]} for r in pass1
    ]
    pass2 = au.compute_appends(register, combined, today, None, None)
    assert pass2 == []


def test_compute_appends_author_override(tmp_path: Path):
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/22 - new event.",
    }]
    new_rows = au.compute_appends(register, [], today, None, "Override Person")
    assert len(new_rows) == 1
    assert new_rows[0]["author"] == "Override Person"


def test_compute_appends_year_override(tmp_path: Path):
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/22 - event.",
    }]
    new_rows = au.compute_appends(register, [], today, year_override=2024,
                                  author_override=None)
    assert len(new_rows) == 1
    assert new_rows[0]["update_date"] == date(2024, 5, 22)
    assert new_rows[0]["update_year"] == 2024


def test_compute_appends_minor_edit_does_not_duplicate(tmp_path: Path):
    """Whitespace/case/punctuation drift on the same (risk_id, date) does NOT
    create a duplicate row - the normalized prefix fingerprint absorbs it."""
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/10 - First Event ",
    }]
    existing = [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2026, 5, 10), "update_year": 2026,
        "author": "Anton Benedict", "note": "first event.",
    }]
    new_rows = au.compute_appends(register, existing, today, None, None)
    assert new_rows == []


def test_compute_appends_same_md_different_year_different_text_appends():
    """An existing Updates row at 2025-09-12 must NOT prevent a NEW log entry
    on 9/12 with DIFFERENT wording from being appended at the inferred year
    (2026). The dedupe is exact (date + note prefix); only identical text on
    the same M/D collapses across years."""
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "9/12 - one-year followup; remediation verified.",
    }]
    existing = [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2025, 9, 12), "update_year": 2025,
        "author": "Anton Benedict",
        "note": "obstruction struck during pier excavation.",
    }]
    new_rows = au.compute_appends(register, existing, today, None, None)
    assert len(new_rows) == 1
    assert new_rows[0]["update_date"] == date(2026, 9, 12)
    assert new_rows[0]["update_year"] == 2026


def test_compute_appends_cold_start_calibration():
    """Phase 2 regen anchored historical entries to 2025. This script's simple
    rule infers 2026 for a 9/12 log entry today. Bootstrap calibration must
    look up historical years for (rid, M, D) and find the existing 2025 row
    by fingerprint, so the cold-start dry-run does not propose duplicates."""
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "9/12 - obstruction struck during pier excavation.",
    }]
    existing = [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2025, 9, 12), "update_year": 2025,
        "author": "Anton Benedict",
        "note": "obstruction struck during pier excavation.",
    }]
    assert au.compute_appends(register, existing, today, None, None) == []


def test_compute_appends_multiple_entries_same_day_all_append():
    """Two log entries on the same M/D with different notes must both append
    as separate Updates rows with the same inferred year."""
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/22 - morning inspection. 5/22 - afternoon followup.",
    }]
    new_rows = au.compute_appends(register, [], today, None, None)
    assert len(new_rows) == 2
    assert new_rows[0]["update_date"] == date(2026, 5, 22)
    assert new_rows[1]["update_date"] == date(2026, 5, 22)
    assert new_rows[0]["update_year"] == new_rows[1]["update_year"] == 2026
    assert new_rows[0]["note"] != new_rows[1]["note"]
    assert new_rows[0]["update_id"] + 1 == new_rows[1]["update_id"]


def test_compute_appends_jan_dec_rollover():
    """Running in January, a 12/X log entry resolves to the prior year (the
    naive year-of-append + 6mo-future rule handles the Jan/Dec rollover)."""
    today = date(2026, 1, 5)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "12/20 - year-end closeout note.",
    }]
    new_rows = au.compute_appends(register, [], today, None, None)
    assert len(new_rows) == 1
    assert new_rows[0]["update_date"] == date(2025, 12, 20)
    assert new_rows[0]["update_year"] == 2025


def test_compute_appends_skips_invalid_date_silently(caplog):
    """A log entry that parses to a real M/D but invalid calendar date
    (e.g. Feb 30) is logged-and-skipped, not crashed."""
    today = date(2026, 5, 23)
    register = [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "2/30 - impossible date entry. 5/22 - valid.",
    }]
    with caplog.at_level(logging.WARNING):
        new_rows = au.compute_appends(register, [], today, None, None)
    assert len(new_rows) == 1
    assert new_rows[0]["update_date"] == date(2026, 5, 22)


# ---------------------------------------------------------------- end-to-end CLI


def test_dry_run_does_not_write(tmp_path: Path, capsys):
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "updates.xlsx"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/22 - new event.",
    }])
    make_updates_xlsx(upd, [])
    upd_mtime_before = upd.stat().st_mtime_ns

    rc = au.main([
        "--register", str(reg), "--updates", str(upd),
        "--today", "2026-05-23", "--dry-run",
    ])
    assert rc == 0

    # File untouched
    assert upd.stat().st_mtime_ns == upd_mtime_before
    wb = load_workbook(upd)
    ws = wb["Risk_Updates"]
    assert ws.max_row == 1  # header only

    out = capsys.readouterr().out
    assert "Rows to append:        1" in out
    assert "--dry-run: no file written; no archive." in out


def test_real_write_appends_rows(tmp_path: Path):
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "updates.xlsx"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/10 - first. 5/22 - new event.",
    }])
    make_updates_xlsx(upd, [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2026, 5, 10), "update_year": 2026,
        "author": "Anton Benedict", "note": "first.",
    }])

    rc = au.main([
        "--register", str(reg), "--updates", str(upd),
        "--today", "2026-05-23",
    ])
    assert rc == 0

    rows = au.load_updates(upd)
    assert len(rows) == 2
    last = rows[-1]
    assert last["risk_id"] == "TONN-CON.01"
    assert au.to_date(last["update_date"]) == date(2026, 5, 22)
    assert last["update_id"] == 2
    assert last["note"] == "new event."


def test_archive_dated_by_last_update_date(tmp_path: Path):
    """Archive filename uses MAX(update_date) in the existing MASTER, not the
    current run datetime. The archive sits in the archive folder, not next to
    the MASTER."""
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "Tonnelle_Risk_Updates_MASTER.xlsx"
    arch = tmp_path / "archive"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/10 - old. 5/22 - new event.",
    }])
    make_updates_xlsx(upd, [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2026, 5, 10), "update_year": 2026,
        "author": "Anton Benedict", "note": "old.",
    }])

    rc = au.main([
        "--register", str(reg), "--updates", str(upd),
        "--archive-dir", str(arch),
        "--today", "2026-05-23",
    ])
    assert rc == 0

    archives = list(arch.glob("*.xlsx"))
    assert len(archives) == 1
    # Filename pattern: Tonnelle_Risk_Updates_<YYMMDD>.xlsx where YYMMDD is
    # MAX(update_date) in pre-write file = 2026-05-10 -> 260510
    assert archives[0].name == "Tonnelle_Risk_Updates_260510.xlsx"
    # Archive is the pre-write snapshot (1 row before this append)
    wb_arch = load_workbook(archives[0])
    assert wb_arch["Risk_Updates"].max_row == 2  # header + 1 row


def test_archive_same_day_disambiguates_with_hhmmss(tmp_path: Path):
    """Second append on the same day finds the existing dated archive and
    falls back to <name>_<YYMMDD>_<HHMMSS>.xlsx."""
    upd = tmp_path / "Tonnelle_Risk_Updates_MASTER.xlsx"
    arch = tmp_path / "archive"
    arch.mkdir()
    # Pre-existing archive that matches the date we're about to produce
    (arch / "Tonnelle_Risk_Updates_260523.xlsx").write_bytes(b"placeholder")

    make_updates_xlsx(upd, [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2026, 5, 23), "update_year": 2026,
        "author": "Anton Benedict", "note": "today.",
    }])

    out = au.archive_previous_master(upd, arch, run_dt=__import__("datetime").datetime(2026, 5, 23, 14, 32, 17))
    assert out is not None
    assert out.name == "Tonnelle_Risk_Updates_260523_143217.xlsx"


def test_dry_run_does_not_archive(tmp_path: Path):
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "Tonnelle_Risk_Updates_MASTER.xlsx"
    arch = tmp_path / "archive"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/22 - new event.",
    }])
    make_updates_xlsx(upd, [])

    rc = au.main([
        "--register", str(reg), "--updates", str(upd),
        "--archive-dir", str(arch),
        "--today", "2026-05-23", "--dry-run",
    ])
    assert rc == 0
    assert not arch.exists() or not any(arch.iterdir())


def test_no_archive_flag_skips_archive(tmp_path: Path):
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "Tonnelle_Risk_Updates_MASTER.xlsx"
    arch = tmp_path / "archive"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/22 - new event.",
    }])
    make_updates_xlsx(upd, [])

    rc = au.main([
        "--register", str(reg), "--updates", str(upd),
        "--archive-dir", str(arch),
        "--today", "2026-05-23", "--no-archive",
    ])
    assert rc == 0
    assert not arch.exists() or not any(arch.iterdir())
    # But the write still happens
    rows = au.load_updates(upd)
    assert len(rows) == 1


def test_no_changes_returns_zero_and_writes_nothing(tmp_path: Path, capsys):
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "updates.xlsx"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/10 - already logged.",
    }])
    make_updates_xlsx(upd, [{
        "update_id": 1, "risk_id": "TONN-CON.01",
        "update_date": date(2026, 5, 10), "update_year": 2026,
        "author": "Anton Benedict", "note": "already logged.",
    }])
    upd_mtime_before = upd.stat().st_mtime_ns

    rc = au.main([
        "--register", str(reg), "--updates", str(upd),
        "--today", "2026-05-23",
    ])
    assert rc == 0
    assert upd.stat().st_mtime_ns == upd_mtime_before  # not rewritten
    out = capsys.readouterr().out
    assert "Nothing to append" in out


def test_discrepancy_flag_surfaces_backdated_entry(tmp_path: Path, capsys):
    """A log entry parsed to a date >45 days from today gets a [FLAG] marker
    and shows up in the DISCREPANCY summary block."""
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "updates.xlsx"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        # 3/15 is 69 days before 5/23 - past the 45-day window
        "mitigation_log": "3/15 - backdated entry RM forgot earlier. 5/22 - recent.",
    }])
    make_updates_xlsx(upd, [])

    rc = au.main([
        "--register", str(reg), "--updates", str(upd),
        "--today", "2026-05-23", "--dry-run",
    ])
    assert rc == 0
    out = capsys.readouterr().out
    # Recent entry not flagged
    assert "2026-05-22" in out
    # Backdated entry flagged
    assert "2026-03-15" in out
    assert "[FLAG]" in out
    assert "DISCREPANCY: 1 row(s)" in out
    assert "(-69d)" in out


def test_discrepancy_flag_skipped_when_within_window(tmp_path: Path, capsys):
    """Entries within ±45 days of today print without [FLAG]."""
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "updates.xlsx"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "5/22 - normal entry.",
    }])
    make_updates_xlsx(upd, [])

    au.main([
        "--register", str(reg), "--updates", str(upd),
        "--today", "2026-05-23", "--dry-run",
    ])
    out = capsys.readouterr().out
    assert "[FLAG]" not in out
    assert "DISCREPANCY" not in out


def test_malformed_in_workbook_does_not_crash(tmp_path: Path, capsys, caplog):
    reg = tmp_path / "register.xlsx"
    upd = tmp_path / "updates.xlsx"
    make_register_xlsx(reg, [{
        "risk_id": "TONN-CON.01",
        "risk_coordinator": "Anton Benedict",
        "mitigation_log": "13/45 - junk. 5/22 - good entry.",
    }])
    make_updates_xlsx(upd, [])

    with caplog.at_level(logging.WARNING):
        rc = au.main([
            "--register", str(reg), "--updates", str(upd),
            "--today", "2026-05-23", "--dry-run",
        ])
    assert rc == 0
    out = capsys.readouterr().out
    assert "Rows to append:        1" in out
    assert any("13/45" in r.message or "out-of-range" in r.message
               for r in caplog.records)
