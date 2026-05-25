"""Phase 1 discovery utility. Reads both xlsx files and prints everything
needed to write docs/01_audit.md. Read-only: no file writes, no mutation
of source data.

Run from project root:
    python scripts/audit_inspect.py
"""
from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
REGISTER_PATH = ROOT / "source_data" / "Tonnelle_Risk_Register_260519.xlsx"
UPDATES_PATH  = ROOT / "source_data" / "Tonnelle_Risk_Updates_MASTER.xlsx"

LOCKED_CATEGORIES = {"Construction", "Field Condition", "Design Change",
                     "Safety", "Environmental", "Political"}
LOCKED_ENTITIES   = {"GDC", "CM", "Contractor", "Shared"}
REQUIRED_FIELDS   = ["status", "risk_coordinator",
                     "probability_score", "cost_impact_score", "schedule_impact_score"]

TODAY = date(2026, 5, 22)  # per CLAUDE.md currentDate

# Reasonable next_review_date window: 1y back, 2y forward
DATE_LO = date(TODAY.year - 1, TODAY.month, TODAY.day)
DATE_HI = date(TODAY.year + 2, TODAY.month, TODAY.day)


# ---------------------------- helpers ----------------------------

def hdr(title: str) -> None:
    print()
    print("=" * 72)
    print(title)
    print("=" * 72)


def sub(title: str) -> None:
    print()
    print(f"-- {title}")


def rows_as_dicts(ws, header_row: int = 1):
    """Yield each data row as a dict keyed by header name."""
    headers = [c.value for c in ws[header_row]]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue
        yield dict(zip(headers, row)), headers


def to_date(v):
    """openpyxl returns Python datetime/date for date-typed cells; passthrough.
    If it returns a number, treat as Excel serial."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel 1900 epoch with leap-bug; openpyxl handles via from_excel
        from openpyxl.utils.datetime import from_excel
        d = from_excel(v)
        return d.date() if isinstance(d, datetime) else d
    return None


# ---------------------------- REGISTER ----------------------------

def audit_register():
    hdr("REGISTER  (data_only=True so formulas return cached values)")
    wb = load_workbook(REGISTER_PATH, data_only=True)
    print(f"sheets: {wb.sheetnames}")
    ws = wb["Risk_Register"]
    print(f"Risk_Register dims: {ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")

    headers = [c.value for c in ws[1]]
    print(f"columns ({len(headers)}): {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    print(f"populated data rows: {len(rows)}")

    if rows:
        sub("sample row 0 (verbatim)")
        for k, v in rows[0].items():
            disp = v if not isinstance(v, str) or len(v) <= 100 else v[:100] + "...[truncated]"
            print(f"   {k!r}: {disp!r}")

    # --- Score recomputation
    sub("score recomputation:  stored vs P * MAX(C, S)")
    mismatches_score = []
    for r in rows:
        p = r.get("probability_score")
        c = r.get("cost_impact_score")
        s = r.get("schedule_impact_score")
        stored = r.get("risk_score_overall")
        try:
            expected = p * max(c, s) if None not in (p, c, s) else None
        except TypeError:
            expected = None
        if expected is not None and stored != expected:
            mismatches_score.append((r.get("risk_id"), p, c, s, stored, expected))
    print(f"   rows with score mismatch: {len(mismatches_score)}")
    for m in mismatches_score:
        print(f"     {m[0]}  P={m[1]} C={m[2]} S={m[3]}  stored={m[4]}  expected={m[5]}")

    # --- Band recomputation
    sub("risk_level band recomputation:  High>=15, Medium 8-14, Low 1-7")
    def expected_band(score):
        if score is None:
            return None
        if score >= 15:
            return "High"
        if score >= 8:
            return "Medium"
        return "Low"
    mismatches_band = []
    for r in rows:
        stored = r.get("risk_level")
        score = r.get("risk_score_overall")
        exp = expected_band(score)
        if exp is not None and stored != exp:
            mismatches_band.append((r.get("risk_id"), score, stored, exp))
    print(f"   rows with band mismatch: {len(mismatches_band)}")
    for m in mismatches_band:
        print(f"     {m[0]}  score={m[1]}  stored={m[2]}  expected={m[3]}")

    # --- Distributions
    sub("distribution: risk_level")
    for k, v in sorted(Counter(r.get("risk_level") for r in rows).items(),
                       key=lambda kv: (kv[0] is None, str(kv[0]))):
        print(f"   {k!r}: {v}")

    sub("distribution: risk_category")
    cat_counter = Counter(r.get("risk_category") for r in rows)
    for k, v in sorted(cat_counter.items(), key=lambda kv: (-kv[1], str(kv[0]))):
        print(f"   {k!r}: {v}")

    sub("distribution: risk_entity")
    ent_counter = Counter(r.get("risk_entity") for r in rows)
    for k, v in sorted(ent_counter.items(), key=lambda kv: (-kv[1], str(kv[0]))):
        print(f"   {k!r}: {v}")

    sub("distribution: status")
    for k, v in sorted(Counter(r.get("status") for r in rows).items(),
                       key=lambda kv: (-kv[1], str(kv[0]))):
        print(f"   {k!r}: {v}")

    sub("distribution: risk_coordinator")
    for k, v in sorted(Counter(r.get("risk_coordinator") for r in rows).items(),
                       key=lambda kv: (-kv[1], str(kv[0]))):
        print(f"   {k!r}: {v}")

    # --- Lookups tab reconciliation
    sub("LOOKUPS tab inspection")
    ws_l = wb["Lookups"]
    print(f"   Lookups dims: {ws_l.dimensions}  max_row={ws_l.max_row}  max_col={ws_l.max_column}")
    print(f"   header row: {[c.value for c in ws_l[1]]}")
    print("   full Lookups content (non-empty cells):")
    for row in ws_l.iter_rows(values_only=True):
        if any(v not in (None, "") for v in row):
            print(f"     {row}")

    # Try to extract category/entity/status lists from Lookups tab.
    # Strategy: any column whose header contains 'categor' / 'entit' / 'status'.
    lookups_cols = defaultdict(list)
    headers_l = [c.value for c in ws_l[1]]
    for row in ws_l.iter_rows(min_row=2, values_only=True):
        for h, v in zip(headers_l, row):
            if h and v not in (None, ""):
                lookups_cols[h].append(v)

    def pick_col(*needles):
        for h, vals in lookups_cols.items():
            if any(n.lower() in str(h).lower() for n in needles):
                return h, vals
        return None, []

    cat_h, cat_l = pick_col("categor")
    ent_h, ent_l = pick_col("entit")
    print(f"   Lookups[{cat_h!r}] = {cat_l}")
    print(f"   Lookups[{ent_h!r}] = {ent_l}")

    # --- Reconciliation
    sub("category reconciliation:  DATA vs LOOKUPS vs LOCKED")
    data_cats = set(c for c in cat_counter if c is not None)
    lookup_cats = set(cat_l)
    print(f"   data categories ({len(data_cats)}): {sorted(data_cats)}")
    print(f"   lookups categories ({len(lookup_cats)}): {sorted(lookup_cats)}")
    print(f"   locked categories ({len(LOCKED_CATEGORIES)}): {sorted(LOCKED_CATEGORIES)}")
    print(f"   in DATA not in LOOKUPS: {sorted(data_cats - lookup_cats)}")
    print(f"   in DATA not in LOCKED:  {sorted(data_cats - LOCKED_CATEGORIES)}")
    print(f"   in LOOKUPS not in LOCKED: {sorted(lookup_cats - LOCKED_CATEGORIES)}")
    print(f"   in LOCKED not in DATA:  {sorted(LOCKED_CATEGORIES - data_cats)} (unused locked values)")

    sub("entity reconciliation:  DATA vs LOOKUPS vs LOCKED")
    data_ents = set(e for e in ent_counter if e is not None)
    lookup_ents = set(ent_l)
    print(f"   data entities ({len(data_ents)}): {sorted(data_ents)}")
    print(f"   lookups entities ({len(lookup_ents)}): {sorted(lookup_ents)}")
    print(f"   locked entities ({len(LOCKED_ENTITIES)}): {sorted(LOCKED_ENTITIES)}")
    print(f"   in DATA not in LOOKUPS: {sorted(data_ents - lookup_ents)}")
    print(f"   in DATA not in LOCKED:  {sorted(data_ents - LOCKED_ENTITIES)}")
    print(f"   in LOOKUPS not in LOCKED: {sorted(lookup_ents - LOCKED_ENTITIES)}")
    print(f"   in LOCKED not in DATA:  {sorted(LOCKED_ENTITIES - data_ents)} (unused locked values)")

    # --- Blank required fields
    sub("blank required-field audit")
    for f in REQUIRED_FIELDS:
        bad = [r.get("risk_id") for r in rows if r.get(f) in (None, "")]
        print(f"   {f}: {len(bad)} blank  -> {bad}")

    # --- next_review_date window
    sub("next_review_date sanity")
    dates = []
    none_count = 0
    for r in rows:
        d = to_date(r.get("next_review_date"))
        if d is None:
            none_count += 1
            continue
        dates.append((r.get("risk_id"), d))
    print(f"   present: {len(dates)}  blank: {none_count}")
    if dates:
        print(f"   min: {min(d for _, d in dates)}   max: {max(d for _, d in dates)}")
        outside = [(rid, d) for rid, d in dates if d < DATE_LO or d > DATE_HI]
        print(f"   outside [{DATE_LO} .. {DATE_HI}]: {len(outside)}")
        for rid, d in outside:
            print(f"     {rid}  {d}")

    return rows  # for cross-file step


# ---------------------------- UPDATES ----------------------------

def audit_updates(register_rows):
    hdr("UPDATES")
    wb = load_workbook(UPDATES_PATH, data_only=True)
    print(f"sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    print(f"dims: {ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")
    headers = [c.value for c in ws[1]]
    print(f"columns ({len(headers)}): {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    print(f"populated data rows: {len(rows)}")

    if rows:
        sub("sample row 0 (verbatim)")
        for k, v in rows[0].items():
            disp = v if not isinstance(v, str) or len(v) <= 100 else v[:100] + "...[truncated]"
            print(f"   {k!r}: {disp!r}")

    # --- date range
    sub("update_date range")
    dts = []
    for r in rows:
        d = to_date(r.get("update_date"))
        if d:
            dts.append(d)
    if dts:
        print(f"   min: {min(dts)}   max: {max(dts)}   count: {len(dts)}")

    # --- year distribution
    sub("distribution by update_year (stored column)")
    yr_counter = Counter(r.get("update_year") for r in rows)
    for k, v in sorted(yr_counter.items(), key=lambda kv: (kv[0] is None, kv[0])):
        print(f"   {k!r}: {v}")

    sub("derived year from update_date (cross-check)")
    der_yr = Counter(d.year for d in dts)
    for k, v in sorted(der_yr.items()):
        print(f"   {k!r}: {v}")

    # --- orphans
    sub("orphan risk_ids (in Updates but not in Register)")
    reg_ids = {r.get("risk_id") for r in register_rows}
    upd_ids = {r.get("risk_id") for r in rows}
    orphans = [r for r in rows if r.get("risk_id") not in reg_ids]
    orphan_ids = sorted({r.get("risk_id") for r in orphans})
    print(f"   distinct orphan risk_ids: {len(orphan_ids)}  -> {orphan_ids}")

    sub("zero-coverage risk_ids (in Register, no Updates row)")
    zero = sorted(reg_ids - upd_ids)
    print(f"   {len(zero)}  -> {zero}")

    return rows


# ---------------------------- CROSS ----------------------------

DATE_LEADER = re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})\s*-\s*", re.MULTILINE)


def parse_mitigation_log(text: str):
    """Return list of (m, d, body) tuples extracted from a mitigation_log cell."""
    if not text or not isinstance(text, str):
        return []
    matches = list(DATE_LEADER.finditer(text))
    out = []
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[start:end].strip()
        out.append((int(m.group("m")), int(m.group("d")), body))
    return out


def normalize(s: str) -> str:
    """Whitespace + case + trailing-punctuation normalization."""
    return re.sub(r"\s+", " ", (s or "")).strip().lower().rstrip(".,;:!?")


def cross_reconcile(register_rows, updates_rows):
    hdr("CROSS:  Register.mitigation_log  vs  Updates.note")
    # Index updates by (risk_id, m, d)
    upd_idx = defaultdict(list)
    for r in updates_rows:
        d = to_date(r.get("update_date"))
        if d:
            upd_idx[(r.get("risk_id"), d.month, d.day)].append(r)

    log_total = 0
    log_present = 0
    pair_match = 0
    pair_drift = 0
    pair_logonly = 0  # mitigation_log entry has no Updates row
    pair_updonly = 0  # Updates row has no mitigation_log entry on that day
    drift_examples = []
    logonly_examples = []
    updonly_examples = []

    for reg in register_rows:
        rid = reg.get("risk_id")
        if not rid:
            continue
        entries = parse_mitigation_log(reg.get("mitigation_log"))
        if entries:
            log_present += 1
        log_total += len(entries)

        # mitigation_log -> Updates direction
        for m, d, body in entries:
            ups = upd_idx.get((rid, m, d), [])
            if not ups:
                pair_logonly += 1
                if len(logonly_examples) < 5:
                    logonly_examples.append((rid, m, d, body[:80]))
                continue
            # Find any matching by normalized text
            if any(normalize(u.get("note", "")) == normalize(body) for u in ups):
                pair_match += 1
            else:
                pair_drift += 1
                if len(drift_examples) < 8:
                    drift_examples.append((rid, m, d, body[:80], (ups[0].get("note") or "")[:80]))

        # Updates direction: for this risk, are there Updates rows on days not in log?
        log_days = {(m, d) for m, d, _ in entries}
        for u in [u for u in updates_rows if u.get("risk_id") == rid]:
            du = to_date(u.get("update_date"))
            if du and (du.month, du.day) not in log_days:
                pair_updonly += 1
                if len(updonly_examples) < 5:
                    updonly_examples.append((rid, du, (u.get("note") or "")[:80]))

    print(f"   risks with any mitigation_log dated entry: {log_present}")
    print(f"   total mitigation_log dated entries parsed: {log_total}")
    print(f"   matched (same risk, same M/D, same note text): {pair_match}")
    print(f"   text drift (same risk+date, different text):   {pair_drift}")
    print(f"   log-only (no Updates row for that date):        {pair_logonly}")
    print(f"   Updates-only (no mitigation_log entry that day):{pair_updonly}")

    sub("drift examples (first 8)")
    for rid, m, d, b, u in drift_examples:
        print(f"     {rid}  {m}/{d}")
        print(f"       log: {b!r}")
        print(f"       upd: {u!r}")

    sub("log-only examples (first 5)")
    for rid, m, d, b in logonly_examples:
        print(f"     {rid}  {m}/{d}  log: {b!r}")

    sub("Updates-only examples (first 5)")
    for rid, dt, n in updonly_examples:
        print(f"     {rid}  {dt}  upd: {n!r}")


# ---------------------------- MAIN ----------------------------

def main():
    if not REGISTER_PATH.exists():
        print(f"missing: {REGISTER_PATH}", file=sys.stderr)
        sys.exit(2)
    if not UPDATES_PATH.exists():
        print(f"missing: {UPDATES_PATH}", file=sys.stderr)
        sys.exit(2)

    reg_rows = audit_register()
    upd_rows = audit_updates(reg_rows)
    cross_reconcile(reg_rows, upd_rows)

    hdr("DONE")


if __name__ == "__main__":
    main()
