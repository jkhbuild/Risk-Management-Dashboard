"""
Phase 5 self-verification helper.

Compute the expected baseline value of each _Measures.Counts / _Measures.Scores
measure against the Risk_Register source workbook. Print as a Markdown table so
the output can be pasted directly into docs/05_semantic_model.md.

Run from project root:
    PYTHONUTF8=1 PYTHONIOENCODING=utf-8 python scripts/verify_phase5_measures.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


REGISTER_PATH = Path("source_data/Tonnelle_Risk_Register_MASTER.xlsx")


def load_register() -> pd.DataFrame:
    wb = load_workbook(REGISTER_PATH, data_only=True, read_only=True)
    ws = wb["Risk_Register"]
    rows = list(ws.iter_rows(values_only=True))
    header, *body = rows
    df = pd.DataFrame(body, columns=header)
    df = df.dropna(subset=["risk_id"])
    return df


def main() -> None:
    df = load_register()

    total_risks = len(df)
    high_risks = (df["risk_level"] == "High").sum()
    medium_risks = (df["risk_level"] == "Medium").sum()
    low_risks = (df["risk_level"] == "Low").sum()

    avg_overall = df["risk_score_overall"].mean()
    avg_cost = df["risk_score_cost"].mean()
    avg_schedule = df["risk_score_schedule"].mean()
    max_overall = df["risk_score_overall"].max()

    rows = [
        ("Total Risks", total_risks, "COUNTROWS(Risk_Register) where risk_id is not null"),
        ("High Risks", int(high_risks), "Rows where risk_level = 'High'"),
        ("Medium Risks", int(medium_risks), "Rows where risk_level = 'Medium'"),
        ("Low Risks", int(low_risks), "Rows where risk_level = 'Low'"),
        ("Avg Risk Score Overall", round(avg_overall, 2), "Mean of risk_score_overall (37 rows)"),
        ("Avg Cost Score", round(avg_cost, 2), "Mean of risk_score_cost (37 rows)"),
        ("Avg Schedule Score", round(avg_schedule, 2), "Mean of risk_score_schedule (37 rows)"),
        ("Max Risk Score", int(max_overall), "Max of risk_score_overall (37 rows)"),
    ]

    print(f"Source: {REGISTER_PATH}")
    print(f"Rows loaded (post-dropna on risk_id): {len(df)}")
    print()
    print("| Measure | Expected value | Source-data computation |")
    print("|---|---|---|")
    for name, value, computation in rows:
        print(f"| `[{name}]` | {value} | {computation} |")

    print()
    print("Row-level integrity:")
    band_counts = df["risk_level"].value_counts(dropna=False).to_dict()
    print(f"  risk_level value counts: {band_counts}")
    print(f"  Sum of bands: {high_risks + medium_risks + low_risks}, total rows: {total_risks}")
    if high_risks + medium_risks + low_risks != total_risks:
        print("  WARNING: band sum != total — investigate stray risk_level values")


if __name__ == "__main__":
    main()
