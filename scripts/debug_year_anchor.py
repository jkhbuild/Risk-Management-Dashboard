"""Inspect every risk's mitigation_log entries with their parsed (m, d) and the
year my regen algorithm assigned. Flag any risk whose first entry's month is
< 6 (i.e., the algorithm would assign 2025 to what may actually be 2026)."""
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
REGISTER = ROOT / "source_data" / "Tonnelle_Risk_Register_260519.xlsx"

DATE_LEADER = re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})\s*-\s*", re.MULTILINE)
PROJECT_START_YEAR = 2025


def parse(text):
    if not text:
        return []
    ms = list(DATE_LEADER.finditer(text))
    out = []
    for i, m in enumerate(ms):
        start = m.end()
        end = ms[i+1].start() if i+1 < len(ms) else len(text)
        out.append((int(m.group("m")), int(m.group("d")), text[start:end].strip()))
    return out


wb = load_workbook(REGISTER, data_only=True)
ws = wb["Risk_Register"]
hdrs = [c.value for c in ws[1]]
rows = [dict(zip(hdrs, r)) for r in ws.iter_rows(min_row=2, values_only=True)
        if any(v is not None for v in r)]

print(f"{'risk_id':<14} {'first_M/D':<10} {'flag'}")
print("-" * 80)
for r in rows:
    entries = parse(r.get("mitigation_log") or "")
    if not entries:
        continue
    first_m, first_d, _ = entries[0]
    months = [e[0] for e in entries]
    flag = ""
    if first_m < 6:
        flag = "FIRST ENTRY MONTH<6 - may need 2026 anchor"
    # Walk algorithm
    year = PROJECT_START_YEAR
    prev = None
    dates = []
    for m, d, _ in entries:
        if prev is not None and m < prev:
            year += 1
        dates.append(f"{year}-{m:02d}-{d:02d}")
        prev = m
    print(f"{r.get('risk_id'):<14} {first_m}/{first_d:<7} {flag}")
    print(f"               months in order: {months}")
    print(f"               assigned dates:  {dates}")
    print()
