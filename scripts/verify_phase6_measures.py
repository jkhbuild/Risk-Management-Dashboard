"""
Phase 6 self-verification helper.

Compute expected baseline values for the _Measures.TimeIntel measures:
  - [Updates Count] grouped by YearMonth (Risk_Updates per month)
  - [Days Since Last Update] per risk_id (today - max update_date)

Also report:
  - min/max update_date (verifies dim_Date coverage seed)
  - count of risks without any updates (Days Since Last Update returns BLANK)

Run from project root:
    PYTHONUTF8=1 PYTHONIOENCODING=utf-8 python scripts/verify_phase6_measures.py
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


UPDATES_PATH = Path("source_data/Tonnelle_Risk_Updates_MASTER.xlsx")
REGISTER_PATH = Path("source_data/Tonnelle_Risk_Register_MASTER.xlsx")


def load_updates() -> pd.DataFrame:
    wb = load_workbook(UPDATES_PATH, data_only=True, read_only=True)
    ws = wb["Risk_Updates"]
    rows = list(ws.iter_rows(values_only=True))
    header, *body = rows
    df = pd.DataFrame(body, columns=header)
    df = df.dropna(subset=["update_id"])
    df["update_date"] = pd.to_datetime(df["update_date"]).dt.date
    return df


def load_register() -> pd.DataFrame:
    wb = load_workbook(REGISTER_PATH, data_only=True, read_only=True)
    ws = wb["Risk_Register"]
    rows = list(ws.iter_rows(values_only=True))
    header, *body = rows
    df = pd.DataFrame(body, columns=header)
    df = df.dropna(subset=["risk_id"])
    return df


def main() -> None:
    updates = load_updates()
    register = load_register()

    today = date(2026, 5, 23)  # CLAUDE.md currentDate; Power BI TODAY() resolves to system date

    min_date = updates["update_date"].min()
    max_date = updates["update_date"].max()

    print(f"Source: {UPDATES_PATH}")
    print(f"Rows loaded (post-dropna on update_id): {len(updates)}")
    print(f"Risks in Register (post-dropna on risk_id): {len(register)}")
    print()
    print("## Date range coverage")
    print(f"  min(update_date): {min_date}")
    print(f"  max(update_date): {max_date}")
    expected_calendar_start = date(min_date.year, 1, 1)
    expected_calendar_end = date(max_date.year, 12, 31)
    print(f"  Expected dim_Date span (per 03 §a CALENDAR seed): "
          f"{expected_calendar_start} to {expected_calendar_end}")
    print(f"  Span: {(expected_calendar_end - expected_calendar_start).days + 1} days")
    print()

    # Monthly Updates Count (the Page 1 trend chart values)
    updates["year_month"] = updates["update_date"].apply(lambda d: f"{d.year:04d}-{d.month:02d}")
    monthly = updates.groupby("year_month").size().sort_index()

    print("## Expected [Updates Count] by YearMonth")
    print("(this is what the Page 1 Risk Activity Over Time line will plot)")
    print()
    print("| YearMonth | Updates Count |")
    print("|---|---|")
    for ym, cnt in monthly.items():
        print(f"| {ym} | {cnt} |")
    print(f"| **TOTAL** | **{monthly.sum()}** |")
    print()

    # Sanity: months in update range with zero updates (continuous axis will show as 0)
    all_months = pd.period_range(start=str(min_date)[:7], end=str(max_date)[:7], freq="M")
    all_month_strs = [str(p) for p in all_months]
    zero_months = [m for m in all_month_strs if m not in monthly.index]
    print(f"Months in [min, max] with zero updates: {zero_months or 'none'}")
    print()

    # Days Since Last Update per risk (expected per-risk values for Page 2 column)
    print("## [Days Since Last Update] per risk")
    print(f"  TODAY() reference: {today}")
    last_update_by_risk = updates.groupby("risk_id")["update_date"].max()
    register_ids = set(register["risk_id"])
    update_ids = set(last_update_by_risk.index)

    risks_no_updates = register_ids - update_ids
    update_only_risks = update_ids - register_ids

    print(f"  Risks in Register with no Updates rows (Days Since Last Update = BLANK): "
          f"{len(risks_no_updates)}")
    if risks_no_updates:
        print(f"    {sorted(risks_no_updates)}")
    print(f"  Updates risk_ids not in Register (will not appear in Top Risks table): "
          f"{len(update_only_risks)}")
    if update_only_risks:
        print(f"    {sorted(update_only_risks)}")

    days_since = last_update_by_risk.apply(lambda d: (today - d).days)
    summary = days_since.describe()
    print()
    print("  Per-risk days-since-last-update summary (37 Register risks, "
          "those with updates):")
    print(f"    min:    {int(summary['min'])} d")
    print(f"    median: {int(summary['50%'])} d")
    print(f"    max:    {int(summary['max'])} d")
    print(f"    count:  {int(summary['count'])}")
    print()

    # Show top-10 most-stale risks (largest Days Since Last Update)
    top10_stale = days_since.sort_values(ascending=False).head(10)
    print("  Top 10 stalest risks (sample for sanity-checking Page 2 sort):")
    print("  | risk_id | Days Since Last Update |")
    print("  |---|---|")
    for rid, d in top10_stale.items():
        print(f"  | {rid} | {d} |")


if __name__ == "__main__":
    main()
