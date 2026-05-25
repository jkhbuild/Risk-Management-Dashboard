"""append_updates.py - Phase 12 operational tool.

Detects new dated entries in `Risk_Register.mitigation_log` that are not
already represented in `Risk_Updates`, and appends them as new rows to
`Risk_Updates`. Idempotent: re-running with no Register changes appends
nothing.

Workflow:
  1) RM emails a dated copy of the Register (Tonnelle_Risk_Register_<YYMMDD>
     .xlsx). Owner archives the dated file to /archive/ and overwrites
     source_data/Tonnelle_Risk_Register_MASTER.xlsx with the same content.
  2) Owner runs `python scripts/append_updates.py --dry-run` and reviews the
     proposed append rows (plus any [FLAG] discrepancies).
  3) Owner runs `python scripts/append_updates.py` to commit. The existing
     Updates MASTER is auto-archived to /archive/<base>_<YYMMDD>.xlsx where
     YYMMDD = MAX(update_date) in the pre-write file. Pass --no-archive to
     skip (rare).
  4) Power BI refresh picks up the new Risk_Updates rows.

The Phase 1 schema (`update_id, risk_id, update_date, update_year, author,
note`) is preserved; the file remains a single-sheet append-only flat table.

Run:
  PYTHONUTF8=1 PYTHONIOENCODING=utf-8 python scripts/append_updates.py [flags]
"""
from __future__ import annotations

import argparse
import logging
import re
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_REGISTER = ROOT / "source_data" / "Tonnelle_Risk_Register_MASTER.xlsx"
DEFAULT_UPDATES = ROOT / "source_data" / "Tonnelle_Risk_Updates_MASTER.xlsx"
DEFAULT_ARCHIVE_DIR = ROOT / "archive"

DATE_LEADER = re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})\s*-\s*", re.MULTILINE)
SIX_MONTHS_DAYS = 183
NOTE_KEY_LEN = 120
DISCREPANCY_DAYS = 45  # ~1.5 months; appends dated outside this window from today get a [FLAG]

UPDATE_COLS = [
    "update_id", "risk_id", "update_date", "update_year", "author", "note",
    "updated_status",
    "probability_at_update", "cost_impact_at_update", "schedule_impact_at_update",
]

# Score-change parsing patterns. Captures the NEW score value (the post-change
# integer). Matches "prob to 3", "probability raised to 3", "cost from 4 to 2",
# "sched 3->5", etc. Note: PIR scores are 1-5 for prob, 0-5 for cost/sched.
SCORE_PATTERNS = {
    "probability_at_update": re.compile(
        r"\b(?:probability|prob)\b[^.;\n]*?\b(?:to|->|=|now|raised to|reduced to|increased to|decreased to)\s*(\d)\b",
        re.IGNORECASE,
    ),
    "cost_impact_at_update": re.compile(
        r"\bcost(?:\s+impact)?\b[^.;\n]*?\b(?:to|->|=|now|raised to|reduced to|increased to|decreased to)\s*(\d)\b",
        re.IGNORECASE,
    ),
    "schedule_impact_at_update": re.compile(
        r"\b(?:sched(?:ule)?(?:\s+impact)?)\b[^.;\n]*?\b(?:to|->|=|now|raised to|reduced to|increased to|decreased to)\s*(\d)\b",
        re.IGNORECASE,
    ),
}


def parse_score_changes(note: str) -> dict:
    """Return a dict of any score columns whose post-change value is detected
    in the note. Caller is responsible for None-vs-omit semantics.

    Examples (informational; for full test coverage see test_append_updates.py):
        >>> parse_score_changes("Probability raised to 4 after design change")
        {'probability_at_update': 4}
        >>> parse_score_changes("cost from 4 to 2; sched 3 -> 5")
        {'cost_impact_at_update': 2, 'schedule_impact_at_update': 5}
        >>> parse_score_changes("monitoring quarterly, no score change")
        {}
    """
    out = {}
    for col, pat in SCORE_PATTERNS.items():
        m = pat.search(note or "")
        if m:
            try:
                out[col] = int(m.group(1))
            except (ValueError, TypeError):
                pass
    return out

# Status inference (mirrors scripts/backfill_status.py rules). When the note text
# matches a keyword, the new updated_status value is set accordingly; otherwise the
# row inherits the prior status. Terminal statuses (Closed/Realized) stay terminal
# unless a later note clearly promotes to another terminal state.
STATUS_RULES = [
    ("Closed", [
        r"\brisk closed\b", r"\bclosed\b", r"\bresolved\b",
        r"\bno longer applicable\b", r"\bnot applicable\b",
        r"\bcancelled\b", r"\bcanceled\b", r"\bwithdrawn\b",
    ]),
    ("Realized", [
        r"\brealized\b", r"\bmaterialized\b",
        r"\bchange order issued\b", r"\bchange order executed\b",
        r"\bbid item added\b", r"\bincident occurred\b", r"\bevent occurred\b",
    ]),
    ("Monitoring", [
        r"\bmonitoring\b", r"\bwatching\b", r"\bunder review\b", r"\bdormant\b",
    ]),
]
TERMINAL_STATUSES = {"Closed", "Realized"}


def infer_updated_status(note: str, prior: str | None) -> str:
    """Infer the post-update status from the note text and the prior status.

    Mirrors scripts/backfill_status.py rules. Returns prior status (default
    'Open') when no keyword fires. Terminal statuses persist unless promoted.
    """
    text = (note or "").lower()
    for status, patterns in STATUS_RULES:
        for pat in patterns:
            if re.search(pat, text):
                # Promote terminal-to-terminal but otherwise hold the terminal state
                if prior in TERMINAL_STATUSES and status not in TERMINAL_STATUSES:
                    return prior
                return status
    return prior or "Open"

logger = logging.getLogger("append_updates")


def normalize(s: str) -> str:
    """Whitespace + case + trailing-punctuation normalization."""
    return re.sub(r"\s+", " ", (s or "")).strip().lower().rstrip(".,;:!?")


def fingerprint(risk_id: str, d: date, note: str) -> tuple[str, str, str]:
    """Dedupe key: (risk_id, ISO date, first NOTE_KEY_LEN chars of normalized
    note). Exact-date: same risk + same FULL date + same note prefix = same
    event. Distinct entries on the same M/D in different years (legitimate
    follow-ups) and multiple entries on the same day (different notes) both
    append correctly.
    """
    return (risk_id, d.isoformat(), normalize(note)[:NOTE_KEY_LEN])


def parse_log(text: str, source_rid: str | None = None) -> list[tuple[int, int, str]]:
    """Return list of (month, day, body) tuples extracted from a mitigation_log
    cell. Tolerates malformed entries: logs a warning, skips, does not crash.
    """
    if not text or not isinstance(text, str):
        return []
    matches = list(DATE_LEADER.finditer(text))
    out: list[tuple[int, int, str]] = []
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[start:end].strip()
        try:
            month = int(m.group("m"))
            day = int(m.group("d"))
            if not (1 <= month <= 12 and 1 <= day <= 31):
                raise ValueError(f"out-of-range M/D: {month}/{day}")
            if not body:
                raise ValueError("empty body after date leader")
            out.append((month, day, body))
        except (ValueError, TypeError) as e:
            logger.warning(
                "Skipping malformed entry%s near offset %d: %s | leader=%r body=%r",
                f" in {source_rid}" if source_rid else "",
                m.start(), e, m.group(0), body[:60],
            )
    return out


def infer_year(month: int, day: int, today: date, override: int | None) -> int | None:
    """Return year for (month, day) given today's date.

    Rule:
      - If --year-override is set, use it (validated against month/day).
      - Otherwise default to today.year; if the resulting date is more than
        ~6 months in the future, roll back to year-1.
    Returns None if (year, month, day) is not a valid calendar date.
    """
    if override is not None:
        try:
            date(override, month, day)
            return override
        except ValueError as e:
            logger.warning("--year-override %d invalid for %d/%d: %s",
                           override, month, day, e)
            return None
    year = today.year
    try:
        candidate = date(year, month, day)
    except ValueError as e:
        logger.warning("Skipping entry: invalid date %d/%d/%d: %s",
                       month, day, year, e)
        return None
    if (candidate - today).days > SIX_MONTHS_DAYS:
        return year - 1
    return year


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def load_register(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Risk_Register"]
    headers = [c.value for c in ws[1]]
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        d = dict(zip(headers, r))
        if not d.get("risk_id"):
            continue
        rows.append(d)
    return rows


def load_updates(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def compute_appends(
    register_rows: list[dict],
    existing_updates: list[dict],
    today: date,
    year_override: int | None,
    author_override: str | None,
    status_override: str | None = None,
    prob_override: int | None = None,
    cost_override: int | None = None,
    sched_override: int | None = None,
) -> list[dict]:
    """Return the list of rows to append. Does not mutate inputs."""
    existing_fp: set[tuple[str, str, str]] = set()
    historical_years: dict[tuple[str, int, int], set[int]] = defaultdict(set)
    max_id = 0
    # Track latest updated_status per risk (chronologically) for inferring new rows
    latest_status_by_risk: dict[str, tuple[date, str]] = {}
    for u in existing_updates:
        d = to_date(u.get("update_date"))
        if d is None:
            continue
        rid_u = u.get("risk_id")
        existing_fp.add(fingerprint(rid_u, d, u.get("note") or ""))
        historical_years[(rid_u, d.month, d.day)].add(d.year)
        uid = u.get("update_id")
        if isinstance(uid, int) and uid > max_id:
            max_id = uid
        s = u.get("updated_status")
        if s and (rid_u not in latest_status_by_risk or d >= latest_status_by_risk[rid_u][0]):
            latest_status_by_risk[rid_u] = (d, str(s))

    new_rows: list[dict] = []
    next_id = max_id
    for reg in register_rows:
        rid = reg.get("risk_id")
        coord = reg.get("risk_coordinator")
        author = author_override or coord or "(unassigned)"
        for month, day, body in parse_log(reg.get("mitigation_log"), source_rid=rid):
            year = infer_year(month, day, today, year_override)
            if year is None:
                continue
            try:
                upd_date = date(year, month, day)
            except ValueError as e:
                logger.warning("Skipping %s %d/%d: invalid date for year %d: %s",
                               rid, month, day, year, e)
                continue
            if fingerprint(rid, upd_date, body) in existing_fp:
                continue
            # Bootstrap calibration: this entry was not matched at the
            # inferred year, but the same (risk_id, M, D, note) may already
            # exist at a different historical year (Phase 2 regen used
            # calibration; this script does not). Check each historical year
            # known for this risk's M/D before declaring the entry new.
            represented = False
            for hist_year in historical_years.get((rid, month, day), ()):
                if hist_year == upd_date.year:
                    continue
                try:
                    hist_date = date(hist_year, month, day)
                except ValueError:
                    continue
                if fingerprint(rid, hist_date, body) in existing_fp:
                    represented = True
                    break
            if represented:
                continue
            next_id += 1
            prior_status = latest_status_by_risk.get(rid, (None, None))[1]
            if status_override:
                new_status = status_override
            else:
                new_status = infer_updated_status(body, prior_status)
            score_changes = parse_score_changes(body)
            if prob_override is not None:
                score_changes["probability_at_update"] = prob_override
            if cost_override is not None:
                score_changes["cost_impact_at_update"] = cost_override
            if sched_override is not None:
                score_changes["schedule_impact_at_update"] = sched_override
            new_rows.append({
                "update_id": next_id,
                "risk_id": rid,
                "update_date": upd_date,
                "update_year": upd_date.year,
                "author": author,
                "note": body,
                "updated_status": new_status,
                "probability_at_update": score_changes.get("probability_at_update"),
                "cost_impact_at_update": score_changes.get("cost_impact_at_update"),
                "schedule_impact_at_update": score_changes.get("schedule_impact_at_update"),
            })
            if score_changes:
                logger.info("[SCORE] %s on %s: %s | %s",
                            rid, upd_date.isoformat(),
                            ", ".join(f"{k}={v}" for k,v in score_changes.items()),
                            body[:60])
            if prior_status and new_status != prior_status:
                logger.info("[STATUS] %s on %s: %s -> %s | %s",
                            rid, upd_date.isoformat(), prior_status, new_status, body[:80])
            existing_fp.add(fingerprint(rid, upd_date, body))
            historical_years[(rid, month, day)].add(upd_date.year)
            latest_status_by_risk[rid] = (upd_date, new_status)
    return new_rows


def write_updates(path: Path, existing_updates: list[dict], new_rows: list[dict]) -> None:
    """Rewrite the Updates workbook with existing + appended rows.
    Single sheet `Risk_Updates`, columns per UPDATE_COLS.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk_Updates"
    ws.append(UPDATE_COLS)
    for u in existing_updates:
        ws.append([u.get(c) for c in UPDATE_COLS])
    for r in new_rows:
        ws.append([r[c] for c in UPDATE_COLS])
    for cell in ws["C"][1:]:
        cell.number_format = "yyyy-mm-dd"
    wb.save(path)


def archive_previous_master(
    updates_path: Path,
    archive_dir: Path,
    run_dt: datetime | None = None,
) -> Path | None:
    """Copy the existing Updates MASTER to archive_dir, dated by MAX(update_date)
    in that file (the semantic "last updated" date).

    Filename convention: `<base>_<YYMMDD>.xlsx` where <base> is the MASTER stem
    with the trailing `_MASTER` removed (e.g. `Tonnelle_Risk_Updates_260512.xlsx`).
    If that path already exists (a second append in the same day), `_HHMMSS` is
    appended for disambiguation.

    Returns the archive path written, or None if updates_path doesn't exist.
    """
    if not updates_path.exists():
        return None
    run_dt = run_dt or datetime.now()
    wb = load_workbook(updates_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    dates: list[date] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) >= 3:
            d = to_date(r[2])
            if d:
                dates.append(d)
    last_date = max(dates) if dates else run_dt.date()

    archive_dir.mkdir(parents=True, exist_ok=True)
    base_stem = updates_path.stem
    if base_stem.endswith("_MASTER"):
        base_stem = base_stem[: -len("_MASTER")]
    yymmdd = last_date.strftime("%y%m%d")
    suffix = updates_path.suffix
    candidate = archive_dir / f"{base_stem}_{yymmdd}{suffix}"
    if candidate.exists():
        hhmmss = run_dt.strftime("%H%M%S")
        candidate = archive_dir / f"{base_stem}_{yymmdd}_{hhmmss}{suffix}"
    shutil.copy2(updates_path, candidate)
    return candidate


def print_summary(existing_count: int, new_rows: list[dict], today: date) -> None:
    print(f"Existing Updates rows: {existing_count}")
    print(f"Rows to append:        {len(new_rows)}")
    if not new_rows:
        return
    print()
    print(f"{'update_id':>9}  {'risk_id':<12}  {'date':<10}  "
          f"{'author':<20}  note")
    print("-" * 116)
    flagged: list[tuple[dict, int]] = []
    for r in new_rows:
        note_disp = (r["note"] or "").replace("\n", " ").replace("\r", " ")
        if len(note_disp) > 50:
            note_disp = note_disp[:47] + "..."
        author = (r["author"] or "")[:20]
        days = (r["update_date"] - today).days
        marker = "  [FLAG]" if abs(days) > DISCREPANCY_DAYS else ""
        print(f"{r['update_id']:>9}  {r['risk_id']:<12}  "
              f"{r['update_date'].isoformat()}  "
              f"{author:<20}  {note_disp}{marker}")
        if abs(days) > DISCREPANCY_DAYS:
            flagged.append((r, days))
    if flagged:
        print()
        print(f"DISCREPANCY: {len(flagged)} row(s) dated more than "
              f"{DISCREPANCY_DAYS} days from today ({today.isoformat()}).")
        print("Verify the inferred year is correct for each. Use "
              "--year-override or edit the source mitigation_log entry "
              "if not.")
        for r, days in flagged:
            sign = "+" if days >= 0 else ""
            note_disp = (r["note"] or "").replace("\n", " ").replace("\r", " ")
            if len(note_disp) > 60:
                note_disp = note_disp[:57] + "..."
            print(f"  {r['risk_id']:<12} {r['update_date'].isoformat()} "
                  f"({sign}{days}d): {note_disp}")


def parse_args(argv=None):
    p = argparse.ArgumentParser(
        prog="append_updates.py",
        description=(
            "Append new dated mitigation_log entries from Risk_Register to "
            "Risk_Updates. Idempotent: re-runs append nothing if no changes "
            "exist. Recommended workflow: --dry-run first, then re-run "
            "(auto-archives the previous MASTER to /archive/ unless --no-archive)."
        ),
    )
    p.add_argument(
        "--register", type=Path, default=DEFAULT_REGISTER,
        help=("Risk_Register .xlsx (default: "
              "source_data/Tonnelle_Risk_Register_MASTER.xlsx)"),
    )
    p.add_argument(
        "--updates", type=Path, default=DEFAULT_UPDATES,
        help=("Risk_Updates .xlsx (default: "
              "source_data/Tonnelle_Risk_Updates_MASTER.xlsx)"),
    )
    p.add_argument(
        "--dry-run", action="store_true",
        help="Print proposed appends; do not write; do not archive.",
    )
    p.add_argument(
        "--no-archive", action="store_true",
        help=("Skip the auto-archive step. By default, before writing, the "
              "existing Updates MASTER is copied to <archive>/<base>_<YYMMDD>"
              ".xlsx where YYMMDD = MAX(update_date) in the file. Second "
              "append in the same day appends _HHMMSS."),
    )
    p.add_argument(
        "--archive-dir", type=Path, default=DEFAULT_ARCHIVE_DIR, metavar="DIR",
        help=f"Archive folder (default: {DEFAULT_ARCHIVE_DIR.relative_to(ROOT)})",
    )
    p.add_argument(
        "--year-override", type=int, default=None, metavar="YYYY",
        help=("Force YYYY for ALL parsed entries instead of the "
              "current-year + 6-month-future rule."),
    )
    p.add_argument(
        "--author", default=None, metavar="NAME",
        help=("Override author for ALL appended rows "
              "(default: risk_coordinator from Register)."),
    )
    p.add_argument(
        "--status-override", default=None, metavar="STATUS",
        choices=["Open", "Monitoring", "Realized", "Closed"],
        help=("Force a specific updated_status for ALL appended rows "
              "(default: inferred from note text via keyword rules; "
              "see scripts/backfill_status.py)."),
    )
    p.add_argument(
        "--prob-override", type=int, default=None, metavar="N", choices=[1,2,3,4,5],
        help="Force probability_at_update=N for ALL appended rows (1-5).",
    )
    p.add_argument(
        "--cost-override", type=int, default=None, metavar="N", choices=[0,1,2,3,4,5],
        help="Force cost_impact_at_update=N for ALL appended rows (0-5).",
    )
    p.add_argument(
        "--sched-override", type=int, default=None, metavar="N", choices=[0,1,2,3,4,5],
        help="Force schedule_impact_at_update=N for ALL appended rows (0-5).",
    )
    p.add_argument(
        "--today", default=None, metavar="YYYY-MM-DD",
        help="Override today's date for year inference (testing aid).",
    )
    p.add_argument(
        "-v", "--verbose", action="store_true",
        help="Verbose logging (DEBUG level).",
    )
    return p.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    if args.today:
        try:
            today = datetime.strptime(args.today, "%Y-%m-%d").date()
        except ValueError as e:
            logger.error("--today must be YYYY-MM-DD: %s", e)
            return 2
    else:
        today = date.today()

    if not args.register.exists():
        logger.error("Register not found: %s", args.register)
        return 2
    if not args.updates.exists():
        logger.error("Updates not found: %s", args.updates)
        return 2

    register_rows = load_register(args.register)
    existing_updates = load_updates(args.updates)

    new_rows = compute_appends(
        register_rows, existing_updates, today,
        args.year_override, args.author,
        status_override=args.status_override,
        prob_override=args.prob_override,
        cost_override=args.cost_override,
        sched_override=args.sched_override,
    )

    print_summary(len(existing_updates), new_rows, today)

    if not new_rows:
        print("\nNothing to append; Risk_Updates is in sync.")
        return 0

    if args.dry_run:
        print("\n--dry-run: no file written; no archive.")
        return 0

    if not args.no_archive:
        try:
            archive_path = archive_previous_master(args.updates, args.archive_dir)
        except OSError as e:
            logger.error("Archive failed (%s); aborting write to avoid losing "
                         "the previous MASTER. Re-run with --no-archive to "
                         "force, or fix the archive folder.", e)
            return 2
        if archive_path:
            print(f"\nArchived previous MASTER: {archive_path}")

    write_updates(args.updates, existing_updates, new_rows)
    print(f"\nWrote {len(new_rows)} new row(s) to {args.updates}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
