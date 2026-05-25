"""Regenerate Risk_Updates from current Risk_Register.mitigation_log per Phase
2 §c-Q3 (docs/02_schema_challenge.md, locked in CLAUDE.md as a user-side
prerequisite for Phase 5).

Inputs (read-only):
  source_data/Tonnelle_Risk_Register_260519.xlsx  (sheet: Risk_Register)
  source_data/Tonnelle_Risk_Updates_MASTER.xlsx   (sheet: Risk_Updates)

Output:
  source_data/Tonnelle_Risk_Updates_REGENERATED.xlsx  (new file, not an overwrite)

After review the user manually:
  1. moves the original Tonnelle_Risk_Updates_260519.xlsx to /archive/
  2. renames the regenerated file to Tonnelle_Risk_Updates_MASTER.xlsx (the
     new canonical name; the dated filename is retired going forward).

Re-running this script reads from MASTER and regenerates a fresh REGENERATED
file. Useful if more entries are added to mitigation_log in Excel.

Regeneration rules:
  - Walk each Register row; parse mitigation_log on '(\\d+)/(\\d+) - ' markers.
  - Year disambiguation (two-stage):
      1. CALIBRATE: if any existing Updates row for this risk has the same
         (month, day) as a parsed log entry, anchor that entry's year to
         the existing row's year. This trusts the prior splitter's
         disambiguation where it ran.
      2. WALK: starting from the anchored year (or the heuristic below if
         no anchor exists), bump year by 1 each time the next entry's
         month is strictly less than the previous entry's. Project span
         is Sep 2025 - May 2026, so a within-year month decrease is the
         signal for a year roll.
      Heuristic for risks without any existing Updates row
      (TONN-CON.31-37, the 7 user-added risks): if the first entry's
      month is < 6 anchor at 2026; otherwise anchor at 2025.
  - author defaults to risk_coordinator from Register (matches the 3:1
    ratio of update rows to risks per coordinator in the current file).
  - Preserve any Updates-only events (existing rows whose (month, day)
    does not appear in that risk's parsed log days). Their date / author /
    note are kept verbatim. Phase 1 audit §c identified 6 such events,
    all terminal closings ('risk closed' / 'risk realized').
  - Sort all rows by (update_date, risk_id) and assign sequential
    update_id starting at 1.
  - Columns written, in order: update_id, risk_id, update_date,
    update_year, author, note. Matches the locked Phase 3 §a schema.

Run:
  PYTHONUTF8=1 PYTHONIOENCODING=utf-8 python scripts/regenerate_updates.py
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
REGISTER_PATH = ROOT / "source_data" / "Tonnelle_Risk_Register_260519.xlsx"
UPDATES_PATH = ROOT / "source_data" / "Tonnelle_Risk_Updates_MASTER.xlsx"
OUTPUT_PATH = ROOT / "source_data" / "Tonnelle_Risk_Updates_REGENERATED.xlsx"

DATE_LEADER = re.compile(r"(?P<m>\d{1,2})/(?P<d>\d{1,2})\s*-\s*", re.MULTILINE)


def parse_log(text: str) -> list[tuple[int, int, str]]:
    """Return list of (month, day, body) tuples in source order."""
    if not text or not isinstance(text, str):
        return []
    matches = list(DATE_LEADER.finditer(text))
    out = []
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[start:end].strip()
        out.append((int(m.group("m")), int(m.group("d")), body))
    return out


def assign_years(
    entries: list[tuple[int, int, str]],
    existing_md_to_year: dict[tuple[int, int], int],
) -> list[tuple[date, str]]:
    """Walk entries in order. Calibration: anchor first entry's year from any
    matching existing-Updates row; otherwise heuristic (month<6 -> 2026 else
    2025). Then bump year each time month strictly decreases between entries.
    """
    if not entries:
        return []

    first_m, _first_d, _ = entries[0]
    # 1) Try to anchor first entry from existing Updates
    anchor_year = existing_md_to_year.get((first_m, entries[0][1]))
    if anchor_year is None:
        # 2) Try any later entry that has a calibration hit, then walk back
        for idx, (m, d, _) in enumerate(entries):
            if (m, d) in existing_md_to_year:
                back_year = existing_md_to_year[(m, d)]
                prev_m = m
                for j in range(idx - 1, -1, -1):
                    em, _ed, _ = entries[j]
                    if em > prev_m:
                        back_year -= 1
                    prev_m = em
                anchor_year = back_year
                break
    if anchor_year is None:
        # 3) Heuristic for risks with no existing Updates calibration
        anchor_year = 2026 if first_m < 6 else 2025

    out = []
    year = anchor_year
    prev_month = None
    for m, d, body in entries:
        if prev_month is not None and m < prev_month:
            year += 1
        out.append((date(year, m, d), body))
        prev_month = m
    return out


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def load_register():
    wb = load_workbook(REGISTER_PATH, data_only=True)
    ws = wb["Risk_Register"]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def load_existing_updates():
    wb = load_workbook(UPDATES_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def main() -> int:
    if not REGISTER_PATH.exists():
        print(f"missing: {REGISTER_PATH}", file=sys.stderr)
        return 2
    if not UPDATES_PATH.exists():
        print(f"missing: {UPDATES_PATH}", file=sys.stderr)
        return 2

    register = load_register()
    existing = load_existing_updates()

    # Index existing Updates by risk_id -> list of rows
    existing_by_risk: dict[str, list[dict]] = defaultdict(list)
    for u in existing:
        existing_by_risk[u.get("risk_id")].append(u)

    new_rows: list[dict] = []
    carryforward_rows: list[dict] = []
    risks_with_log = 0
    risks_without_coordinator = []

    for reg in register:
        rid = reg.get("risk_id")
        coord = reg.get("risk_coordinator")
        if not coord:
            risks_without_coordinator.append(rid)

        entries = parse_log(reg.get("mitigation_log"))
        if entries:
            risks_with_log += 1
        # Build calibration map from existing Updates rows for this risk
        existing_md_to_year: dict[tuple[int, int], int] = {}
        for u in existing_by_risk.get(rid, []):
            du = to_date(u.get("update_date"))
            if du:
                existing_md_to_year[(du.month, du.day)] = du.year
        dated = assign_years(entries, existing_md_to_year)
        log_days = {(d.month, d.day) for d, _ in dated}

        # Mitigation_log-derived rows
        for d, body in dated:
            new_rows.append({
                "risk_id": rid,
                "update_date": d,
                "update_year": d.year,
                "author": coord or "(unassigned)",
                "note": body,
            })

        # Carry forward Updates-only rows for this risk
        for u in existing_by_risk.get(rid, []):
            du = to_date(u.get("update_date"))
            if du is None:
                continue
            if (du.month, du.day) in log_days:
                continue
            carryforward_rows.append({
                "risk_id": rid,
                "update_date": du,
                "update_year": du.year,
                "author": u.get("author") or coord or "(unassigned)",
                "note": u.get("note"),
            })

    all_rows = new_rows + carryforward_rows
    # Sort by update_date, then risk_id, then note (stable tie-break)
    all_rows.sort(key=lambda r: (r["update_date"], r["risk_id"], r["note"] or ""))

    # Assign sequential update_id
    for i, r in enumerate(all_rows, start=1):
        r["update_id"] = i

    # ---------- write output ----------
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Risk_Updates"
    cols = ["update_id", "risk_id", "update_date", "update_year", "author", "note"]
    ws_out.append(cols)
    for r in all_rows:
        ws_out.append([r[c] for c in cols])
    # Format date column
    for cell in ws_out["C"][1:]:
        cell.number_format = "yyyy-mm-dd"

    wb_out.save(OUTPUT_PATH)

    # ---------- summary ----------
    print("=" * 72)
    print("Risk_Updates regeneration summary")
    print("=" * 72)
    print(f"Output: {OUTPUT_PATH}")
    print(f"Register rows scanned: {len(register)}")
    print(f"  with at least one dated log entry: {risks_with_log}")
    if risks_without_coordinator:
        print(f"  WARNING risks lacking coordinator: {risks_without_coordinator}")
    print(f"Mitigation_log-derived rows: {len(new_rows)}")
    print(f"Updates-only carryforward rows: {len(carryforward_rows)}")
    print(f"Total output rows: {len(all_rows)}")
    if all_rows:
        print(f"Date range: {min(r['update_date'] for r in all_rows)}"
              f" .. {max(r['update_date'] for r in all_rows)}")
    print()
    print("Carryforward rows (preserved Updates-only events):")
    for r in carryforward_rows:
        print(f"  {r['update_date']}  {r['risk_id']}  {r['author']}  "
              f"{(r['note'] or '')[:80]!r}")

    # Author distribution sanity check
    from collections import Counter
    print()
    print("Author distribution in regenerated output:")
    for a, n in sorted(Counter(r["author"] for r in all_rows).items(),
                       key=lambda kv: (-kv[1], kv[0])):
        print(f"  {a}: {n}")

    # Coverage check: did any risk end up with no rows?
    by_risk_out = Counter(r["risk_id"] for r in all_rows)
    zero = [reg.get("risk_id") for reg in register
            if by_risk_out.get(reg.get("risk_id"), 0) == 0]
    if zero:
        print(f"\nWARNING risks with zero updates in output: {zero}")
    else:
        print("\nCoverage: every Register risk has at least one update row.")

    # Per-risk date trace (review aid)
    print("\nPer-risk date assignment (review for any obviously wrong year):")
    by_risk_dates: dict[str, list] = defaultdict(list)
    for r in all_rows:
        by_risk_dates[r["risk_id"]].append(r["update_date"])
    for rid in sorted(by_risk_dates.keys()):
        dts = sorted(by_risk_dates[rid])
        print(f"  {rid}: {[d.isoformat() for d in dts]}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
